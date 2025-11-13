#!/usr/bin/env python3
"""
Generate fake Excel data for testing TrueStreamingMultiSheetProcessor
Creates 2 files:
- test_3sheets.xlsx: 3 sheets (HOP_DONG: 150 rows, CIF: 120 rows, TAP: 100 rows)
- test_2sheets.xlsx: 2 sheets (HOP_DONG: 180 rows, CIF: 150 rows)
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import random
import string

# Helper functions
def random_date(start_year=2020, end_year=2024):
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    return start + timedelta(days=random.randint(0, (end - start).days))

def random_contract_number():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))

def random_cif():
    return ''.join(random.choices(string.digits, k=9))

def random_name():
    first_names = ['NGUYEN', 'TRAN', 'LE', 'PHAM', 'HOANG', 'PHAN', 'VU', 'VO', 'DANG', 'BUI']
    middle_names = ['VAN', 'THI', 'DUC', 'MINH', 'HONG', 'THANH', 'QUOC', 'ANH']
    last_names = ['AN', 'BINH', 'CUONG', 'DUNG', 'HAI', 'HOA', 'LINH', 'MAI', 'NAM', 'PHONG']
    return f"{random.choice(first_names)} {random.choice(middle_names)} {random.choice(last_names)}"

def random_box_code():
    return 'BOX_' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

# Shared data
kho_vpbank = ['HCM', 'HN', 'DN', 'CT']
ma_dv = ['DV001', 'DV002', 'DV003', 'DV004']
tnbg = ['NGUYEN VAN A', 'TRAN THI B', 'LE VAN C']
loai_hs_list = ['LD', 'MD', 'CC', 'OD', 'TTK', 'HDHM', 'TSBD', 'KSSV', 'Bao thanh toán', 'Biên nhận thế chấp']
loai_hs_cif = ['PASS TTN', 'SCF VEERFIN', 'Trình cấp TD không qua CPC', 'Hồ sơ mở TKTT nhưng không giải ngân']
luong_hs = ['HSTD thường', 'HSTD nhanh']
phan_han_td = ['Vĩnh viễn', 'Ngắn hạn', 'Trung hạn', 'Dài hạn']
san_pham = ['Cho vay tiêu dùng', 'Cho vay mua nhà', 'Thẻ tín dụng', 'Dịch vụ thanh toán']
trang_thai_pdm = ['Hoàn thành', 'Đang xử lý', 'Chờ phê duyệt']

def create_hop_dong_sheet(wb, row_count=150):
    """Create HSBG_theo_hop_dong sheet"""
    ws = wb.create_sheet('HSBG_theo_hop_dong')
    headers = [
        'Kho VPBank', 'Mã đơn vị', 'Trách nhiệm bàn giao', 'Số hợp đồng', 'Tên tập', 'Số lượng tập',
        'Số CIF/ CCCD/ CMT khách hàng', 'Tên khách hàng', 'Phân khúc khách hàng', 'Ngày phải bàn giao',
        'Ngày bàn giao', 'Ngày giải ngân', 'Ngày đến hạn', 'Loại hồ sơ', 'Luồng hồ sơ',
        'Phân hạn cấp TD', 'Ngày dự kiến tiêu hủy', 'Sản phẩm', 'Trạng thái case PDM', 'Ghi chú',
        'Mã thùng', 'Ngày nhập kho VPBank', 'Ngày chuyển kho Crown', 'Khu vực', 'Hàng', 'Cột',
        'Tình trạng thùng', 'Trạng thái thùng', 'Thời hạn cấp TD', 'Mã DAO', 'Mã TS', 'RRT.ID', 'Mã NQ'
    ]
    ws.append(headers)

    # Style required columns
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    required_cols = [0, 1, 2, 3, 6, 7, 11, 13, 14, 15, 17, 20]
    for col in required_cols:
        ws.cell(1, col + 1).fill = blue_fill

    # Generate rows
    for i in range(row_count):
        giai_ngan_date = random_date(2020, 2023)
        den_han_date = giai_ngan_date + timedelta(days=random.randint(365, 3650))
        phan_han = random.choice(phan_han_td)

        # Calculate expected destruction date
        if phan_han == 'Vĩnh viễn':
            du_kien_tieu_huy = datetime(9999, 12, 31)
        elif phan_han == 'Ngắn hạn':
            du_kien_tieu_huy = giai_ngan_date + timedelta(days=365*5)
        elif phan_han == 'Trung hạn':
            du_kien_tieu_huy = giai_ngan_date + timedelta(days=365*10)
        else:  # Dài hạn
            du_kien_tieu_huy = giai_ngan_date + timedelta(days=365*15)

        # Calculate thoi_han_cap_td
        months_diff = (den_han_date.year - giai_ngan_date.year) * 12 + (den_han_date.month - giai_ngan_date.month)
        thoi_han_cap_td = max(1, months_diff)

        row = [
            random.choice(kho_vpbank),
            random.choice(ma_dv),
            random.choice(tnbg),
            random_contract_number(),
            f'TAP_{i+1:03d}' if random.random() > 0.3 else '',
            random.randint(1, 5) if random.random() > 0.3 else '',
            random_cif(),
            random_name(),
            random.choice(['Retail', 'SME', 'Corporate']),
            random_date(2020, 2024),
            random_date(2020, 2024),
            giai_ngan_date,
            den_han_date,
            random.choice(loai_hs_list),
            random.choice(luong_hs),
            phan_han,
            du_kien_tieu_huy,
            random.choice(san_pham),
            random.choice(trang_thai_pdm),
            '',
            random_box_code(),
            random_date(2020, 2024),
            random_date(2020, 2024),
            f'KV{random.randint(1,5)}',
            f'H{random.randint(1,20)}',
            f'C{random.randint(1,30)}',
            random.choice(['Tốt', 'Khá', 'Trung bình']),
            random.choice(['Đang lưu', 'Đã xuất', 'Chờ tiêu hủy']),
            thoi_han_cap_td,
            f'DAO{random.randint(1000,9999)}',
            f'TS{random.randint(1000,9999)}',
            f'RRT{random.randint(10000,99999)}',
            f'NQ{random.randint(1000,9999)}'
        ]
        ws.append(row)

    return row_count

def create_cif_sheet(wb, row_count=120):
    """Create HSBG_theo_CIF sheet"""
    ws = wb.create_sheet('HSBG_theo_CIF')
    headers = [
        'Kho VPBank', 'Mã đơn vị', 'Trách nhiệm bàn giao', 'Số CIF khách hàng', 'Tên khách hàng',
        'Tên tập', 'Số lượng tập', 'Phân khúc khách hàng', 'Ngày phải bàn giao', 'Ngày bàn giao',
        'Ngày giải ngân', 'Loại hồ sơ', 'Luồng hồ sơ', 'Phân hạn cấp TD', 'Sản phẩm',
        'Trạng thái case PDM', 'Ghi chú', 'Mã NQ', 'Mã thùng', 'Ngày nhập kho VPBank',
        'Ngày chuyển kho Crown', 'Khu vực', 'Hàng', 'Cột', 'Tình trạng thùng', 'Trạng thái thùng'
    ]
    ws.append(headers)

    # Style
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    required_cols = [0, 1, 2, 3, 4, 10, 11, 12, 13, 14, 18]
    for col in required_cols:
        ws.cell(1, col + 1).fill = blue_fill

    # Generate rows
    for i in range(row_count):
        row = [
            random.choice(kho_vpbank),
            random.choice(ma_dv),
            random.choice(tnbg),
            random_cif(),
            random_name(),
            f'TAP_CIF_{i+1:03d}' if random.random() > 0.3 else '',
            random.randint(1, 5) if random.random() > 0.3 else '',
            random.choice(['Retail', 'SME', 'Corporate']),
            random_date(2020, 2024),
            random_date(2020, 2024),
            random_date(2020, 2023),
            random.choice(loai_hs_cif),
            'HSTD thường',
            'Vĩnh viễn',
            random.choice(san_pham),
            random.choice(trang_thai_pdm),
            '',
            f'NQ{random.randint(1000,9999)}',
            random_box_code(),
            random_date(2020, 2024),
            random_date(2020, 2024),
            f'KV{random.randint(1,5)}',
            f'H{random.randint(1,20)}',
            f'C{random.randint(1,30)}',
            random.choice(['Tốt', 'Khá', 'Trung bình']),
            random.choice(['Đang lưu', 'Đã xuất', 'Chờ tiêu hủy'])
        ]
        ws.append(row)

    return row_count

def create_tap_sheet(wb, row_count=100):
    """Create HSBG_theo_tap sheet"""
    ws = wb.create_sheet('HSBG_theo_tap')
    headers = [
        'Kho VPBank', 'Mã đơn vị', 'Trách nhiệm bàn giao', 'Tháng phát sinh', 'Tên tập', 'Số lượng tập',
        'Ngày phải bàn giao', 'Ngày bàn giao', 'Loại hồ sơ', 'Luồng hồ sơ', 'Phân hạn cấp TD',
        'Ngày dự kiến tiêu hủy', 'Sản phẩm', 'Trạng thái case PDM', 'Ghi chú', 'Mã thùng',
        'Ngày nhập kho VPBank', 'Ngày chuyển kho Crown', 'Khu vực', 'Hàng', 'Cột',
        'Tình trạng thùng', 'Trạng thái thùng'
    ]
    ws.append(headers)

    # Style
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    required_cols = [0, 1, 2, 3, 8, 9, 10, 11, 12, 15]
    for col in required_cols:
        ws.cell(1, col + 1).fill = blue_fill

    # Generate rows
    for i in range(row_count):
        thang_ps_date = random_date(2020, 2024)
        thang_ps = thang_ps_date.strftime('%m/%Y')

        row = [
            random.choice(kho_vpbank),
            random.choice(ma_dv),
            random.choice(tnbg),
            thang_ps,
            f'TAP_KSSV_{i+1:03d}',
            random.randint(1, 10),
            random_date(2020, 2024),
            random_date(2020, 2024),
            'KSSV',
            'HSTD thường',
            'Vĩnh viễn',
            datetime(9999, 12, 31),
            'KSSV',
            random.choice(trang_thai_pdm),
            '',
            random_box_code(),
            random_date(2020, 2024),
            random_date(2020, 2024),
            f'KV{random.randint(1,5)}',
            f'H{random.randint(1,20)}',
            f'C{random.randint(1,30)}',
            random.choice(['Tốt', 'Khá', 'Trung bình']),
            random.choice(['Đang lưu', 'Đã xuất', 'Chờ tiêu hủy'])
        ]
        ws.append(row)

    return row_count

# =================== Generate File 1: 3 Sheets ===================
print("Generating test_3sheets.xlsx...")
wb1 = openpyxl.Workbook()
wb1.remove(wb1.active)

rows_hop_dong = create_hop_dong_sheet(wb1, 150)
rows_cif = create_cif_sheet(wb1, 120)
rows_tap = create_tap_sheet(wb1, 100)

wb1.save('test_3sheets.xlsx')
print(f"[OK] test_3sheets.xlsx created: HOP_DONG({rows_hop_dong}), CIF({rows_cif}), TAP({rows_tap})")

# =================== Generate File 2: 2 Sheets ===================
print("\nGenerating test_2sheets.xlsx...")
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)

rows_hop_dong2 = create_hop_dong_sheet(wb2, 180)
rows_cif2 = create_cif_sheet(wb2, 150)

wb2.save('test_2sheets.xlsx')
print(f"[OK] test_2sheets.xlsx created: HOP_DONG({rows_hop_dong2}), CIF({rows_cif2})")

print("\n[SUCCESS] Both files generated successfully!")
print("Files location: d:\\project\\neobrutalism-crm\\")
